# -*- coding: utf-8 -*-
from fastapi import FastAPI, File, UploadFile, Form
from fastapi.responses import JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
import pandas as pd
import io
import re
import base64
import os
import json
import time
import hashlib
from datetime import datetime, timedelta, date, time as dt_time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import google.generativeai as genai
from PIL import Image
import fitz  # PyMuPDF para processar PDFs
from dotenv import load_dotenv

# Carrega variáveis de ambiente do arquivo .env
load_dotenv()

# --- CONFIGURAÇÃO E CONSTANTES GLOBAIS ---
app = FastAPI(title="API Conversora de Ponto - PontoSync v2.1")

origins = [
    "https://cartaodeponto.netlify.app",
    "http://127.0.0.1:5500",
    "http://localhost:5500",
    "http://localhost:3000",
    "*",  # Allow all origins for direct API calls
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["GET", "POST"],
    allow_headers=["*"],
)

# Verificação da API Key do Gemini - SUPORTE A MÚLTIPLAS CHAVES
GEMINI_API_KEYS = []
GEMINI_MODELS = []  # Lista de modelos instanciados

# Carrega chaves do .env (suporta GEMINI_API_KEY e GEMINI_API_KEY_2)
key1 = os.getenv("GEMINI_API_KEY")
key2 = os.getenv("GEMINI_API_KEY_2")

if key1:
    GEMINI_API_KEYS.append(key1)
if key2:
    GEMINI_API_KEYS.append(key2)

if not GEMINI_API_KEYS:
    print("    ⚠️ AVISO: Nenhuma GEMINI_API_KEY encontrada no .env - Processamento de PDF/Fotos desabilitado")
else:
    # Instancia um modelo para cada chave
    for idx, api_key in enumerate(GEMINI_API_KEYS, 1):
        try:
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-2.5-flash')
            GEMINI_MODELS.append({'key': api_key, 'model': model, 'key_index': idx})
            print(f"    ✅ Gemini API Key #{idx} configurada com sucesso")
        except Exception as e:
            print(f"    ❌ Erro ao configurar API Key #{idx}: {e}")
    
    if GEMINI_MODELS:
        print(f"    🔑 Total de {len(GEMINI_MODELS)} chave(s) API disponível(is) com fallback automático")

DIAS_SEMANA = {
    0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira', 3: 'Quinta-feira',
    4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'
}

# --- MODELOS PYDANTIC (Para Recalcular) ---
class DiaRegistro(BaseModel):
    data: str
    dia_semana: str
    batidas: str
    batidas_4cols: Optional[dict] = None  # NOVO v4.0: 4 slots independentes {ent1, sai1, ent2, sai2}
    saldo: str
    status: str
    alerta: bool
    total: Optional[str] = None  # NOVO: Total trabalhado formatado
    noturno_base: Optional[float] = None  # NOVO: Minutos reais noturno

class FuncionarioPreview(BaseModel):
    funcionario: str
    normais: str
    dever: str
    extras_comuns: str
    extras_100: str
    saldo: str
    dias: List[DiaRegistro]

class RecalcularPayload(BaseModel):
    dados_corrigidos: dict  # O JSON completo que estava no front
    configuracoes: dict     # settings do usuário
    
    class Config:
        extra = 'allow'  # Permite campos extras para robustez

# ===== CATÁLOGO DE JORNADAS CLT (IMUTÁVEL) =====
# Dicionário centralizado com todas as escalas brasileiras suportadas
CATALOGO_JORNADAS_CLT = {
    'clt_5x2_padrao': {
        'nome': '5x2 Padrão (44h)',
        'tipo': 'Semanal',
        'descricao': 'Segunda a sexta: 8h/dia, Sábado: 4h, Domingo: Folga',
        'meta_semana_minutos': 2640,
        'metas_por_dia': {
            0: 480,   # Segunda: 8h
            1: 480,   # Terça: 8h
            2: 480,   # Quarta: 8h
            3: 480,   # Quinta: 8h
            4: 480,   # Sexta: 8h
            5: 240,   # Sábado: 4h (ÚTIL)
            6: 0      # Domingo: Folga
        },
        'sabado_util_automatico': True  # NOVO: Sábado é ÚTIL automaticamente
    },
    'clt_5x2_comp': {
        'nome': '5x2 Compensado (44h)',
        'tipo': 'Semanal',
        'descricao': 'Segunda a sexta: 8h48/dia (528min), Sábado/Domingo: Folga',
        'meta_semana_minutos': 2640,
        'metas_por_dia': {
            0: 528,   # Segunda: 8h48
            1: 528,   # Terça: 8h48
            2: 528,   # Quarta: 8h48
            3: 528,   # Quinta: 8h48
            4: 528,   # Sexta: 8h48
            5: 0,     # Sábado: Folga
            6: 0      # Domingo: Folga
        },
        'sabado_util_automatico': False  # Sábado é FOLGA
    },
    'clt_6x1_com': {
        'nome': '6x1 Comércio (44h)',
        'tipo': 'Semanal',
        'descricao': 'Segunda a sábado: 7h20/dia (440min), Domingo: Folga DSR',
        'meta_semana_minutos': 2640,
        'metas_por_dia': {
            0: 440,   # Segunda: 7h20
            1: 440,   # Terça: 7h20
            2: 440,   # Quarta: 7h20
            3: 440,   # Quinta: 7h20
            4: 440,   # Sexta: 7h20
            5: 440,   # Sábado: 7h20 (ÚTIL)
            6: 0      # Domingo: DSR
        },
        'sabado_util_automatico': True  # OBRIGATÓRIO para comércio
    },
    'clt_5x1': {
        'nome': '5x1 (Ciclo)',
        'tipo': 'Cíclico',
        'descricao': '5 dias trabalho, 1 dia folga. Meta 8h/dia',
        'meta_semana_minutos': 2400,  # 5 dias × 8h
        'metas_por_dia': {
            0: 480,   # Seg-Sex: 8h cada
            1: 480,
            2: 480,
            3: 480,
            4: 480,
            5: 0,     # Sábado: Folga (variável)
            6: 0      # Domingo: Folga (variável)
        },
        'sabado_util_automatico': False  # Depende do ciclo
    },
    'clt_12x36': {
        'nome': '12x36 Plantão',
        'tipo': 'Cíclico',
        'descricao': 'Plantão 12h (720min) seguido de 36h de folga. Ciclo: 48h',
        'meta_semana_minutos': 2640,
        'metas_por_dia': {
            'ciclo': [720, 0, 0]
        },
        'sabado_util_automatico': False
    },
    'estagio_6h': {
        'nome': 'Estágio 6h',
        'tipo': 'Semanal',
        'descricao': 'Estágio: Segunda a sexta, 6h/dia. Sábado/Domingo: Folga',
        'meta_semana_minutos': 1800,  # 30 horas
        'metas_por_dia': {
            0: 360,   # Segunda: 6h
            1: 360,   # Terça: 6h
            2: 360,   # Quarta: 6h
            3: 360,   # Quinta: 6h
            4: 360,   # Sexta: 6h
            5: 0,     # Sábado: Folga
            6: 0      # Domingo: Folga
        },
        'sabado_util_automatico': False
    },
    'clt_personalizada': {
        'nome': 'Personalizada',
        'tipo': 'Dinâmico',
        'descricao': 'Jornada customizada via settings do usuário',
        'meta_semana_minutos': None,
        'metas_por_dia': None,
        'sabado_util_automatico': False  # Usuário define
    }
}

# ===== FUNÇÕES DE CÁLCULO CLT SÊNIOR =====

def calcular_adicional_noturno_estrito(inicio: datetime, fim: datetime) -> int:
    """
    NOVO v4.2: Calcula APENAS os minutos que caem na janela [22:00 às 05:00].
    
    Regra: Interseção entre o horário trabalhado E o período noturno CLT.
    
    Súmula 155 TST:
    - SE jornada COMEÇOU à noite (22:00+) E terminou APÓS 05:00 → bônus continua
    - SE jornada COMEÇOU de dia → bônus para às 05:00
    
    Retorna: Minutos reais (sem redução) que caem em [22:00-05:00]
    """
    PERIODO_NOTURNO_INICIO = 22  # 22:00
    PERIODO_NOTURNO_FIM = 5      # 05:00
    
    # Ajusta para cross-day ANTES de calcular horas
    if fim < inicio:
        fim = fim + timedelta(days=1)
    
    # Obtém horários em horas decimais
    inicio_hora = inicio.hour + inicio.minute / 60.0
    fim_hora = fim.hour + fim.minute / 60.0
    
    minutos_noturno = 0.0
    
    # CASO 1: Entrada à noite (22:00+)
    if inicio_hora >= PERIODO_NOTURNO_INICIO:
        # Calcula da entrada até 05:00 do dia seguinte
        proxima_05 = inicio.replace(hour=5, minute=0, second=0, microsecond=0)
        if proxima_05 <= inicio:
            proxima_05 += timedelta(days=1)
        
        # Fim está antes de 05:00?
        fim_efetivo = min(fim, proxima_05)
        minutos_noturno = (fim_efetivo - inicio).total_seconds() / 60.0
    
    # CASO 2: Entrada de madrugada (00:00-05:00)
    elif inicio_hora < PERIODO_NOTURNO_FIM and inicio_hora >= 0:
        # Calcula do início até 05:00
        proxima_05 = inicio.replace(hour=5, minute=0, second=0, microsecond=0)
        minutos_noturno = (proxima_05 - inicio).total_seconds() / 60.0
    
    # CASO 3: Entrada de dia (05:00-22:00) mas termina à noite/madrugada
    elif inicio_hora >= PERIODO_NOTURNO_FIM and inicio_hora < PERIODO_NOTURNO_INICIO:
        # Cruza a janela noturna?
        if fim_hora >= PERIODO_NOTURNO_INICIO or fim_hora < PERIODO_NOTURNO_FIM:
            # Sim, cruza - calcular desde 22:00 até fim
            noite_inicio = inicio.replace(hour=PERIODO_NOTURNO_INICIO, minute=0, second=0, microsecond=0)
            if noite_inicio <= inicio:
                noite_inicio += timedelta(days=1)
            minutos_noturno = (fim - noite_inicio).total_seconds() / 60.0
    
    return int(round(minutos_noturno))

def calcular_reducao_hora_noturna(inicio: datetime, fim: datetime, noturno_ativo: bool = False) -> tuple:
    """
    REFATORADO v4.5 (Enterprise): Cálculo estrito de hora noturna (Art. 73 CLT).
    
    Regras de Ouro:
    1. Janela Noturna: 22:00 às 05:00.
    2. Redução: 52min30s -> 60min (Fator 1.142857).
    3. Súmula 155 TST (Prorrogação):
       - Se jornada COMEÇA no período noturno (>= 22:00) E estende após as 05:00 -> Aplica noturno na prorrogação.
       - Se jornada COMEÇA de dia/tarde e entra na noite -> Para o noturno às 05:00.
    
    Retorna: (tempo_real_segundos, tempo_reduzido_segundos, minutos_noturno_inteiros)
    """
    if not noturno_ativo:
        duracao_seg = (fim - inicio).total_seconds()
        return (duracao_seg, duracao_seg, 0)
    
    PERIODO_NOTURNO_INICIO = 22.0
    PERIODO_NOTURNO_FIM = 5.0
    FATOR_REDUCAO = 60.0 / 52.5  # 1.142857...
    
    # Ajusta para cross-day
    fim_ajustado = fim if fim >= inicio else fim + timedelta(days=1)
    duracao_total_seg = (fim_ajustado - inicio).total_seconds()
    
    # --- Passo 1: Normalização Decimal ---
    inicio_dec = inicio.hour + (inicio.minute / 60.0)
    fim_dec = fim_ajustado.hour + (fim_ajustado.minute / 60.0) + (24.0 if fim_ajustado.day > inicio.day else 0)
    
    # --- Passo 2: Identificar Janela Noturna Relevante ---
    # A janela noturna padrão é das 22h (22.0) até às 05h do dia seguinte (29.0)
    JANELA_INICIO = 22.0
    JANELA_FIM = 29.0 # 05:00 do dia seguinte
    
    segundos_noturno = 0.0
    
    # Verifica se houve trabalho na janela 22:00 - 05:00
    overlap_inicio = max(inicio_dec, JANELA_INICIO)
    overlap_fim = min(fim_dec, JANELA_FIM)
    
    if overlap_fim > overlap_inicio:
        horas_noturnas = overlap_fim - overlap_inicio
        segundos_noturno += horas_noturnas * 3600.0
        
    # --- Passo 3: Súmula 155 (Prorrogação da Noite) ---
    # Só aplica se a jornada COMEÇOU integralmente dentro ou antes da noite, 
    # mas a regra clássica é: cumpriu integralmente o horário noturno E prorrogou.
    # Simplificação Enterprise: Se iniciou >= 22h, prorroga após as 05h.
    
    if inicio_dec >= JANELA_INICIO and fim_dec > JANELA_FIM:
        # Prorrogação: tudo que excede 05:00 (29.0)
        horas_prorrogacao = fim_dec - JANELA_FIM
        segundos_noturno += horas_prorrogacao * 3600.0
        
    # --- Passo 4: Cálculos Finais ---
    segundos_diurnos = max(0, duracao_total_seg - segundos_noturno)
    segundos_noturno_reduzidos = segundos_noturno * FATOR_REDUCAO
    
    tempo_reduzido_seg = segundos_diurnos + segundos_noturno_reduzidos
    
    return (duracao_total_seg, tempo_reduzido_seg, int(round(segundos_noturno / 60.0)))

def calcular_meta_dinamica_escala(escala_tipo: str, dia_semana_num: int, data_atual_obj: date, 
                                  data_inicio_escala: Optional[date] = None) -> timedelta:
    """
    NOVO v3.0: Calcula a meta diária DINAMICAMENTE baseada no tipo de escala.
    
    Resolve o problema dos 40 minutos: 6x1 tem meta de 440 min (7h20), não 480.
    
    Args:
        escala_tipo: 'clt_5x2_padrao', 'clt_6x1_com', 'clt_12x36', etc.
        dia_semana_num: 0=Seg, ..., 5=Sab, 6=Dom
        data_atual_obj: Data para cálculo (importante para ciclos)
        data_inicio_escala: Data de início do ciclo (para 12x36)
    
    Returns:
        timedelta com a meta para o dia
    """
    
    # Padrão: 5x2
    if escala_tipo == 'clt_5x2_padrao':
        if dia_semana_num == 5:  # Sábado
            return timedelta(hours=4)
        elif dia_semana_num == 6:  # Domingo
            return timedelta(hours=0)
        else:  # Seg-Sex
            return timedelta(hours=8)
    
    # 6x1 Comércio: 440 min (7h20) Seg-Sab, 0 Domingo
    elif escala_tipo == 'clt_6x1_com':
        if dia_semana_num == 6:  # Domingo (DSR)
            return timedelta(minutes=0)
        else:  # Seg-Sab
            return timedelta(minutes=440)  # 7h20
    
    # 5x2 Compensado: 528 min (8h48) Seg-Sex, 0 Sab/Dom
    elif escala_tipo == 'clt_5x2_comp':
        if dia_semana_num in [5, 6]:  # Sab/Dom (Folga)
            return timedelta(minutes=0)
        else:  # Seg-Sex
            return timedelta(minutes=528)  # 8h48
    
    # 12x36 Plantão: Ciclo de 2 dias (1 trabalha, 1 folga)
    elif escala_tipo == 'clt_12x36' and data_inicio_escala:
        try:
            if isinstance(data_inicio_escala, str):
                if '/' in data_inicio_escala:
                    data_init = datetime.strptime(data_inicio_escala, '%d/%m/%Y').date()
                else:
                    data_init = datetime.strptime(data_inicio_escala, '%Y-%m-%d').date()
            else:
                data_init = data_inicio_escala if isinstance(data_inicio_escala, date) else data_inicio_escala.date()
            
            # Calcula diferença em dias a partir da data inicial
            diff_dias = (data_atual_obj - data_init).days
            
            # Dia par (0, 2, 4...) = trabalha 12h, dia ímpar = folga
            if diff_dias % 2 == 0:
                return timedelta(hours=12)
            else:
                return timedelta(hours=0)
        except:
            # Fallback se erro no parse
            return timedelta(hours=8)
    
    # Parcial 30h: 6h Seg-Sex, folga Sab/Dom
    elif escala_tipo == 'clt_parcial_30h':
        if dia_semana_num in [5, 6]:  # Sab/Dom (Folga)
            return timedelta(hours=0)
        else:  # Seg-Sex
            return timedelta(hours=6)
    
    # Personalizada ou desconhecida: padrão 8h
    else:
        if dia_semana_num == 6:
            return timedelta(hours=0)
        else:
            return timedelta(hours=8)

    """
    Implementa PAREAMENTO INTELIGENTE DE BATIDAS com cross-day logic.
    
    Regra: Se a saída ocorrer até `max_intervalo_horas` após a entrada,
    trata-se do mesmo turno, mesmo que cruze a meia-noite.
    
    Retorna lista de tuplas: (entrada_datetime, saida_datetime)
    """
    if len(horarios) < 2:
        return []
    
    pares = []
    i = 0
    
    while i < len(horarios) - 1:
        entrada = horarios[i]
        saida = horarios[i + 1]
        
        # Se a saída é anterior à entrada (cruzou meia-noite), adiciona 1 dia
        if saida < entrada:
            saida = saida + timedelta(days=1)
        
        intervalo = saida - entrada
        
        # Se o intervalo é razoável (entrada e saída), é um par válido
        if intervalo <= timedelta(hours=max_intervalo_horas):
            pares.append((entrada, saida))
            i += 2  # Salta para o próximo par
        else:
            # Entrada sem saída correspondente
            i += 1
    
    return pares

def aplicar_tolerancia_clt(variacao_total_dia_minutos: float, tolerancia_limite: int = 10) -> tuple:
    """
    Implementa tolerância conforme Art. 58, §1º CLT v4.0 (Variação Total do Dia).
    
    Regra NOVA (v4.0 - Market Ready):
    - Calcula a Variação Total do Dia (VTD): soma algébrica de todos atrasos/antecipações
    - Se |VTD| ≤ 10 minutos: ZERO desconto E ZERO abono
    - Se |VTD| > 10 minutos: Computa atraso integralmente (todos os minutos, sem tolerância)
    
    Parâmetros:
        variacao_total_dia_minutos (float): VTD positivo (atraso) ou negativo (antecipação)
        tolerancia_limite (int): Limite em minutos (padrão 10 para CLT)
    
    Retorna: (minutos_abonados, minutos_descontados, observacao)
    
    Exemplos:
    - VTD = 5 min:    (0, 0, "Tolerado")         (dentro de 10min)
    - VTD = -8 min:   (0, 0, "Tolerado")         (antecipação dentro de 10min)
    - VTD = 12 min:   (0, 12, "Integral")        (excedeu 10min, desconta tudo)
    - VTD = -15 min:  (-15, 0, "Antecipação")    (antecipação integral)
    """
    abs_vtd = abs(variacao_total_dia_minutos)
    
    # Dentro da tolerância de 10 minutos
    if abs_vtd <= tolerancia_limite:
        return (0, 0, f"Tolerado (VTD: {variacao_total_dia_minutos:+.0f}min ≤ {tolerancia_limite}min)")
    
    # Excedeu tolerância - aplica regra integral
    if variacao_total_dia_minutos > 0:
        # Atraso (VTD positivo) - desconta integral
        return (0, variacao_total_dia_minutos, f"Desconto integral (VTD: {variacao_total_dia_minutos:.0f}min > {tolerancia_limite}min)")
    else:
        # Antecipação (VTD negativo) - abona integral
        return (abs_vtd, 0, f"Abono integral (VTD: {variacao_total_dia_minutos:.0f}min < -{tolerancia_limite}min)")

# ===== NOVO v6.0: APURAÇÃO SEMANAL DE EXTRAS (44h CLT) =====

def calcular_extras_semanal(dados_semana: dict, jornada_semanal_minutos: int = 2640, debug: bool = True) -> tuple:
    """
    NOVO v6.0: Calcula extras com apuração semanal CLT.
    
    Regras (conforme prática contábil):
    1. Jornada semanal = 44h (2640 min) - configurável
    2. Domingo/Feriado = 100% sempre (prioridade legal)
    3. Excedente real semanal restante = 50%
    4. Compensação entre dias é permitida
    5. Não há dupla contagem
    
    NOTA: Os cálculos seguem a semana ISO, não o mês civil.
    
    Args:
        dados_semana: dict com estrutura por semana ISO
        jornada_semanal_minutos: limite semanal (default 2640 = 44h)
        debug: se True, imprime logs detalhados (desativar em produção)
    
    Returns:
        (total_extras_50, total_extras_100) como timedelta
    """
    jornada_semanal = timedelta(minutes=jornada_semanal_minutos)
    
    total_50 = timedelta(0)
    total_100 = timedelta(0)
    
    for num_semana, dados in dados_semana.items():
        # REGRA 1: 100% = todo trabalho em domingo/feriado (prioridade legal)
        extras_100_semana = dados['horas_dom_fer']
        
        # REGRA 2: Excedente real = total trabalhado - 44h
        excedente = max(timedelta(0), dados['total'] - jornada_semanal)
        
        # REGRA 3: 50% = excedente - já contado como 100% (evita dupla contagem)
        extras_50_semana = max(timedelta(0), excedente - extras_100_semana)
        
        # REGRA 4: Limite - 50% não pode exceder horas úteis trabalhadas
        extras_50_semana = min(extras_50_semana, dados['horas_uteis'])
        
        # LOG para validação (controlado por flag debug)
        if debug:
            total_horas = dados['total'].total_seconds() / 3600
            dom_fer_horas = dados['horas_dom_fer'].total_seconds() / 3600
            e50_horas = extras_50_semana.total_seconds() / 3600
            e100_horas = extras_100_semana.total_seconds() / 3600
            print(f"  📊 Semana {num_semana}: Total={total_horas:.2f}h | Dom/Fer={dom_fer_horas:.2f}h | 50%={e50_horas:.2f}h | 100%={e100_horas:.2f}h")
        
        total_50 += extras_50_semana
        total_100 += extras_100_semana
    
    return total_50, total_100

# --- FUNÇÕES AUXILIARES DE NORMALIZAÇÃO v3.0 (REFATORAÇÃO) ---

def normalizar_horario(valor_entrada: str) -> Optional[dt_time]:
    """
    NOVO v3.0: Parser de horários BLINDADO com múltiplos formatos.
    
    Aceita: "0800", "08:00", "8:00", "800", "8", None, ""
    Retorna: datetime.time ou None
    
    Regra CLT crítica: Se campo vazio ou None, retorna None (NÃO ZERA A LINHA)
    """
    if not valor_entrada or not str(valor_entrada).strip():
        return None
    
    valor = str(valor_entrada).strip()
    
    # Remove espaços e transforma em número
    valor_limpo = valor.replace(':', '').replace(' ', '')
    
    if not valor_limpo or not valor_limpo.isdigit():
        return None
    
    # Processa diferentes comprimentos
    if len(valor_limpo) == 4:
        # "0800" ou "0730" - HHmm com zero à esquerda
        horas = int(valor_limpo[:2])
        minutos = int(valor_limpo[2:])
    elif len(valor_limpo) == 3:
        # "800" ou "730" - Hmmm sem zero à esquerda
        horas = int(valor_limpo[0])
        minutos = int(valor_limpo[1:])
    elif len(valor_limpo) <= 2:
        # "8" ou "08" - Apenas horas
        horas = int(valor_limpo)
        minutos = 0
    else:
        # Mais de 4 dígitos - inválido
        return None
    
    # Validação de intervalos
    if not (0 <= horas <= 23) or not (0 <= minutos <= 59):
        return None
    
    try:
        return dt_time(hour=horas, minute=minutos, second=0)
    except ValueError:
        return None

def format_td(td):
    """Formata timedelta para string HH:MM (Apenas para o Preview visual)"""
    total_seconds = td.total_seconds()
    sign = "-" if total_seconds < 0 else "+"
    total_seconds = abs(total_seconds)
    total_minutes = round(total_seconds / 60)
    hours = int(total_minutes // 60)
    minutes = int(total_minutes % 60)
    
    if hours == 0 and minutes == 0: 
        sign = ""
    elif sign == "+" and (hours > 0 or minutes > 0): 
        sign = "+" 
    return f"{sign}{hours:02d}:{minutes:02d}"

def timedelta_to_excel_time(td: timedelta) -> float:
    """
    Converte timedelta para fração de dia do Excel.
    Excel armazena tempo como float onde 1.0 = 24 horas.
    """
    if td.total_seconds() < 0:
        return 0.0  # Excel não suporta valores negativos de tempo nativamente
    return td.total_seconds() / 86400.0  # 86400 segundos em um dia

def time_to_excel_time(t: dt_time) -> float:
    """
    Converte datetime.time para fração de dia do Excel.
    """
    total_seconds = t.hour * 3600 + t.minute * 60 + t.second
    return total_seconds / 86400.0

def parear_batidas_por_turno(horarios: List[datetime], data_base: date, max_intervalo_horas: int = 12) -> List[tuple]:
    """
    REFATORAÇÃO CIRÚRGICA v5.0: Pareamento com JANELA DE CORTE.
    
    REGRA CRÍTICA: Batidas entre 00:00-05:00 são SAÍDAS do dia anterior,
    não entradas do dia atual. Devem ser empurradas para o final.
    
    Exemplo: [04:01, 10:30, 14:00, 15:00] → [10:30, 14:00, 15:00, 04:01]
    Pareamento correto: (10:30→14:00), (15:00→04:01+1dia)
    """
    if len(horarios) < 2:
        return []
    
    # JANELA DE CORTE: Separa batidas de madrugada (00:00-05:00)
    HORA_CORTE = 5  # Batidas antes das 05:00 são saídas do dia anterior
    
    batidas_madrugada = []
    batidas_normais = []
    
    for h in horarios:
        if h.hour < HORA_CORTE:
            batidas_madrugada.append(h)
        else:
            batidas_normais.append(h)
    
    # Ordena apenas as normais por hora, madrugada vai para o final
    batidas_normais_sorted = sorted(batidas_normais, key=lambda x: x.time())
    
    # Combina: primeiro as diurnas/vespertinas, depois as de madrugada
    horarios_reordenados = batidas_normais_sorted + batidas_madrugada
    
    # Agora pareia: Entrada→Saída→Entrada→Saída
    pares = []
    i = 0
    
    while i < len(horarios_reordenados) - 1:
        entrada = horarios_reordenados[i]
        saida = horarios_reordenados[i + 1]
        
        # Se a saída é de madrugada (hora < 5), adiciona 1 dia para cálculo correto
        if saida.hour < HORA_CORTE and entrada.hour >= HORA_CORTE:
            saida = saida + timedelta(days=1)
        
        intervalo = saida - entrada
        
        # Intervalo válido (até 16h para jornadas noturnas longas)
        if timedelta(0) < intervalo <= timedelta(hours=16):
            pares.append((entrada, saida))
            i += 2
        else:
            # Batida órfã - pula
            i += 1
    
    return pares


def ajustar_horarios_jornada_noturna(horarios: List[datetime], data_base: date, noturno_ativo: bool = False) -> List[datetime]:
    """
    REFATORAÇÃO CIRÚRGICA v5.0: Reorganiza batidas para pareamento correto.
    
    PROIBIDO usar sorted() direto! Isso quebra jornadas noturnas.
    
    Usa a Janela de Corte implementada em parear_batidas_por_turno.
    
    Retorna lista na ordem correta: Entrada1, Saída1, Entrada2, Saída2
    """
    if len(horarios) < 2:
        return horarios
    
    # Usa pareamento inteligente com Janela de Corte
    pares = parear_batidas_por_turno(horarios, data_base)
    
    if not pares:
        # Sem pares válidos - retorna na ordem original
        return horarios
    
    # Reconstrói lista a partir dos pares
    horarios_ajustados = []
    for entrada, saida in pares:
        horarios_ajustados.append(entrada)
        # Remove o +1 dia adicionado para cálculo (mantém hora original)
        saida_original = saida
        if saida.date() > entrada.date():
            # Saída foi ajustada, volta para a hora original (sem mudar data do registro)
            saida_original = datetime.combine(data_base, saida.time())
        horarios_ajustados.append(saida_original)
    
    return horarios_ajustados

def processar_txt(conteudo: str) -> List[dict]:
    """Processa arquivo TXT e retorna lista de registros"""
    dados = []
    padrao_data_hora = re.compile(r'(\d{2}[./-]\d{2}[./-]\d{4})\s+(\d{2}:\d{2}:\d{2})')
    
    for linha in conteudo.splitlines():
        match = padrao_data_hora.search(linha)
        if match:
            try:
                data_str, hora_str = match.groups()
                data_str_norm = data_str.replace('/', '.').replace('-', '.')
                
                info_inicial = linha[:match.start()].split()
                nome = info_inicial[1] if len(info_inicial) > 1 else 'N/A'
                
                dados.append({
                    "nome": nome,
                    "data": datetime.strptime(data_str_norm, "%d.%m.%Y").date(),
                    "hora": datetime.strptime(hora_str, "%H:%M:%S").time()
                })
            except (ValueError, IndexError):
                continue
    
    return dados

def call_gemini_safe(prompt, img):
    """
    Encapsulamento seguro para chamadas Gemini com retry exponencial E FALLBACK DE MÚLTIPLAS CHAVES.
    
    Implementa:
    - Retry com backoff exponencial: [5, 12, 25] segundos
    - Tratamento de erro 429 (quota excedida)
    - NOVO: Fallback automático entre múltiplas chaves API
    - Tratamento de filtros de segurança (Safety filters)
    
    Args:
        prompt: Texto do prompt
        img: Imagem para análise
    
    Returns:
        Response da API Gemini
        
    Raises:
        HTTPException: Em caso de quota excedida em TODAS as chaves ou bloqueio de segurança
    """
    from fastapi import HTTPException
    
    if not GEMINI_MODELS:
        raise HTTPException(
            status_code=503,
            detail="Nenhuma API Key do Gemini configurada. Configure GEMINI_API_KEY no arquivo .env"
        )
    
    retries = [5, 12, 25]  # Intervalos de espera em segundos
    
    # NOVO: Tenta cada chave API disponível
    for model_info in GEMINI_MODELS:
        model = model_info['model']
        key_index = model_info['key_index']
        
        print(f"🔑 Tentando API Key #{key_index}...")
        
        for i, wait_time in enumerate(retries + [0]):
            try:
                # Reconfigura a API com a chave atual antes de cada tentativa
                genai.configure(api_key=model_info['key'])
                response = model.generate_content([prompt, img])
                
                # Sucesso!
                print(f"✅ Sucesso com API Key #{key_index}")
                return response
                
            except Exception as e:
                err_msg = str(e).lower()
                
                # Tratamento: Erro 429 - Quota excedida
                if ("429" in err_msg or "quota" in err_msg or "resource" in err_msg):
                    if i < len(retries):
                        print(f"⏳ Erro 429 na Key #{key_index}. Aguardando {wait_time}s antes da tentativa {i+2}/{len(retries)+1}...")
                        time.sleep(wait_time)
                        continue
                    else:
                        # Esgotou tentativas nesta chave, tenta a próxima
                        print(f"❌ API Key #{key_index} esgotada após {len(retries)+1} tentativas. Tentando próxima chave...")
                        break  # Sai do loop de retries e vai para a próxima chave
                
                # Tratamento: Filtros de segurança da IA
                if "safety" in err_msg or "blocked" in err_msg or "policy" in err_msg:
                    print(f"❌ ERRO CRÍTICO: Conteúdo bloqueado pelos filtros de segurança da IA")
                    raise HTTPException(
                        status_code=400,
                        detail="O arquivo foi bloqueado pelos filtros de segurança da IA. Verifique o conteúdo do documento."
                    )
                
                # Outras exceções: relança sem retry
                print(f"❌ ERRO na Key #{key_index}: {e}")
                raise
    
    # Se chegou aqui, todas as chaves falharam
    print(f"❌ ERRO CRÍTICO: Todas as {len(GEMINI_MODELS)} chave(s) API excederam a cota")
    raise HTTPException(
        status_code=429,
        detail=f"Cota de processamento excedida em todas as {len(GEMINI_MODELS)} chave(s) API. Por favor, aguarde alguns minutos e tente novamente."
    )

def processar_pdf_com_gemini(pdf_bytes: bytes, filename: str) -> List[dict]:
    """Processa PDF usando Gemini Vision para extrair dados do cartão de ponto"""
    if not GEMINI_MODELS:
        raise ValueError("Nenhuma GEMINI_API_KEY configurada. Configure no arquivo .env")
    
    try:
        pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
        dados = []
        
        for page_num in range(len(pdf_document)):
            page = pdf_document[page_num]
            pix = page.get_pixmap()  # Original resolution
            img_bytes = pix.tobytes("png")
            
            # Free pixmap memory immediately
            del pix
            
            img = Image.open(io.BytesIO(img_bytes))
            
            # Free raw bytes after creating PIL image
            del img_bytes
            
            prompt = """Analise este cartão de ponto e extraia TODOS os registros visíveis.

INSTRUÇÕES CRÍTICAS PARA DATAS:
1. O cartão de ponto tem mês/ano no cabeçalho - ENCONTRE E USE essas informações
2. Cada linha tem apenas o DIA (1, 2, 3...) - você DEVE adicionar mês/ano
3. Se o cabeçalho diz "NOVEMBRO/2024", então dia "5" vira "05/11/2024"
4. NUNCA retorne apenas o dia - SEMPRE retorne data completa DD/MM/YYYY

INSTRUÇÕES PARA HORÁRIOS:
1. Ignore completamente a última coluna (assinaturas)
2. Extraia APENAS: Entrada, Saída Almoço, Retorno Almoço, Saída
3. Converta TODOS os horários para formato HH:MM:SS
4. Se um horário estiver ilegível, use null
5. O nome do funcionário está no topo do documento

FORMATO DE RESPOSTA (JSON puro, sem markdown):
{
  "mes": "11",
  "ano": "2024",
  "funcionario": "NOME COMPLETO DO FUNCIONARIO",
  "registros": [
    {
      "data": "01/11/2024",
      "entrada": "07:30:00",
      "saida_almoco": "12:00:00",
      "retorno_almoco": "13:00:00",
      "saida": "17:30:00"
    }
  ]
}

IMPORTANTE: 
- Retorne APENAS o JSON
- Sem texto adicional
- Sem ```json
- Sem explicações
- SEMPRE complete as datas com mês/ano do cabeçalho"""

            print(f"📄 Processando página {page_num + 1} com gemini-1.5-flash...")
            response = call_gemini_safe(prompt, img)
            print(f"✅ Sucesso com gemini-1.5-flash")
            
            # Free PIL image after API call
            img.close()
            del img
            
            if not response:
                continue
            
            texto_resposta = response.text.strip()
            texto_resposta = texto_resposta.replace('```json', '').replace('```', '').strip()
            
            try:
                json_data = json.loads(texto_resposta)
                print(f"[DOC] JSON recebido (página {page_num + 1}): {json_data.get('funcionario', 'N/A')}")
                dados.extend(converter_json_gemini_para_registros(json_data))
            except json.JSONDecodeError as e:
                print(f"[AVISO] Erro ao parsear JSON (página {page_num + 1}): {e}")
                continue
            
            # Force garbage collection after each page
            import gc
            gc.collect()
        
        pdf_document.close()
        return dados
        
    except Exception as e:
        raise ValueError(f"Erro ao processar PDF: {str(e)}")

def processar_imagem_com_gemini(img_bytes: bytes, filename: str) -> List[dict]:
    """Processa imagem (JPG/PNG) usando Gemini Vision"""
    if not GEMINI_MODELS:
        raise ValueError("Nenhuma GEMINI_API_KEY configurada. Configure no arquivo .env")
    
    try:
        img = Image.open(io.BytesIO(img_bytes))
        
        prompt = """Analise este cartão de ponto e extraia TODOS os registros visíveis.

INSTRUÇÕES CRÍTICAS PARA DATAS:
1. O cartão de ponto tem mês/ano no cabeçalho - ENCONTRE E USE essas informações
2. Cada linha tem apenas o DIA (1, 2, 3...) - você DEVE adicionar mês/ano
3. Se o cabeçalho diz "NOVEMBRO/2024", então dia "5" vira "05/11/2024"
4. NUNCA retorne apenas o dia - SEMPRE retorne data completa DD/MM/YYYY

INSTRUÇÕES PARA HORÁRIOS:
1. Ignore completamente a última coluna (assinaturas)
2. Extraia APENAS: Entrada, Saída Almoço, Retorno Almoço, Saída
3. Converta TODOS os horários para formato HH:MM:SS
4. Se um horário estiver ilegível, use null
5. O nome do funcionário está no topo do documento

FORMATO DE RESPOSTA (JSON puro, sem markdown):
{
  "mes": "11",
  "ano": "2024",
  "funcionario": "NOME COMPLETO DO FUNCIONARIO",
  "registros": [
    {
      "data": "01/11/2024",
      "entrada": "07:30:00",
      "saida_almoco": "12:00:00",
      "retorno_almoco": "13:00:00",
      "saida": "17:30:00"
    }
  ]
}

IMPORTANTE: Retorne APENAS o JSON, sem texto adicional"""

        tempos_espera = [5, 10, 20]
        max_tentativas = len(tempos_espera) + 1
        
        print(f"[PROCESS] Processando imagem com gemini-2.5-flash...")
        response = call_gemini_safe(prompt, img)
        print(f"[OK] Sucesso com gemini-2.5-flash")
        
        if not response:
            raise ValueError("Não foi possível obter resposta do Gemini")
        
        texto_resposta = response.text.strip()
        texto_resposta = texto_resposta.replace('```json', '').replace('```', '').strip()
        
        try:
            json_data = json.loads(texto_resposta)
            print(f"[IMG] JSON recebido: {json_data.get('funcionario', 'N/A')}")
            return converter_json_gemini_para_registros(json_data)
        except json.JSONDecodeError as e:
            print(f"[ERRO] Erro ao parsear JSON: {e}")
            raise ValueError(f"Erro ao parsear JSON: {e}")
        
    except Exception as e:
        raise ValueError(f"Erro ao processar imagem: {str(e)}")

def converter_json_gemini_para_registros(json_data: dict) -> List[dict]:
    """Converte JSON do Gemini para o formato esperado pelo sistema"""
    dados = []
    funcionario = json_data.get("funcionario", "N/A")
    
    mes_json = json_data.get("mes")
    ano_json = json_data.get("ano")
    mes_ano = None
    
    if mes_json and ano_json:
        try:
            mes_ano = (int(mes_json), int(ano_json))
            print(f"[DATA] Mês/Ano detectado: {mes_ano[0]:02d}/{mes_ano[1]}")
        except (ValueError, TypeError):
            pass
    
    print(f"\n{'='*70}")
    print(f"[USER] Funcionário: {funcionario}")
    print(f"[LIST] Processando {len(json_data.get('registros', []))} registros")
    print(f"{'='*70}\n")
    
    for idx, registro in enumerate(json_data.get('registros', []), 1):
        try:
            data_str = registro.get("data")
            if not data_str:
                continue
            
            print(f"Registro #{idx}: data='{data_str}'")
            
            data_obj = None
            
            if '/' in data_str or '-' in data_str:
                separador = '/' if '/' in data_str else '-'
                partes = data_str.split(separador)
                
                if len(partes) == 3:
                    try:
                        dia, mes, ano = int(partes[0]), int(partes[1]), int(partes[2])
                        data_obj = date(ano, mes, dia)
                        if not mes_ano:
                            mes_ano = (mes, ano)
                    except (ValueError, IndexError):
                        pass
            
            if not data_obj and data_str.strip().isdigit():
                dia = int(data_str.strip())
                
                if mes_ano:
                    try:
                        data_obj = date(mes_ano[1], mes_ano[0], dia)
                        print(f"   [DATA] Data completada: {dia:02d} -> {data_obj.strftime('%d/%m/%Y')}")
                    except ValueError as e:
                        print(f"   [AVISO] Data inválida: {e}")
                        continue
                else:
                    print(f"   [AVISO] Mês/Ano não identificado")
                    continue
            
            if not data_obj:
                print(f"   [AVISO] Formato não reconhecido: '{data_str}'")
                continue
            
            horarios_encontrados = []
            for campo in ["entrada", "saida_almoco", "retorno_almoco", "saida"]:
                hora_str = registro.get(campo)
                if hora_str and hora_str != "null" and str(hora_str).strip():
                    hora_str = str(hora_str).strip()
                    try:
                        if len(hora_str.split(':')) == 2:
                            hora_str += ":00"
                        hora_obj = datetime.strptime(hora_str, "%H:%M:%S").time()
                        horarios_encontrados.append((campo, hora_obj))
                    except ValueError as e:
                        print(f"   [AVISO] Erro no horário '{campo}={hora_str}': {e}")
                        continue
            
            if horarios_encontrados:
                print(f"   [HORA] Horários: {', '.join([f'{c}={h.strftime('%H:%M')}' for c, h in horarios_encontrados])}")
                
                for campo, hora in horarios_encontrados:
                    dados.append({
                        "nome": funcionario,
                        "data": data_obj,
                        "hora": hora
                    })
            else:
                print(f"   [AVISO] Nenhum horário válido")
                
        except (ValueError, KeyError) as e:
            print(f"   [ERRO] Erro: {e}")
            continue
    
    print(f"\n[OK] Total de {len(dados)} registros criados\n")
    return dados

# ===== FUNÇÃO REFATORADA: LÓGICA DE CÁLCULO COM BATIDAS SEPARADAS =====
def calcular_relatorio(dados_brutos: List[dict], settings: dict, status_overrides: dict = None):
    """
    REFATORADO v4.0 (PontoSync Critical Fix)
    
    ===========================================
    RESPONSABILIDADES (FONTE ÚNICA DE VERDADE):
    ===========================================
    - Calcular tempo trabalhado em minutos/horas
    - Aplicar tolerância CLT (Art. 58)
    - Aplicar redução de hora noturna (Art. 73)
    - Gerar dados para preview e Excel
    
    ===========================================
    NÃO FAZ (deixado para o contador):
    ===========================================
    - Calcular valor financeiro de extras
    - Decidir sobre compensação de banco de horas
    - Interpretar regras de convenção coletiva
    - Decidir juridicamente sobre DSR
    
    ===========================================
    TRADE-OFFS ACEITOS (v4.0):
    ===========================================
    - Saldo é informativo apenas (banco_horas_informativo=True)
    - Percentuais de extra são fixos (50%/100%) e ocultados da UI
    - Frontend não recalcula - apenas renderiza
    
    Defaults quando sem configuração:
    - jornada_minutos: 480 (8h)
    - tolerancia: 10 minutos
    - escala_tipo: clt_5x2_padrao
    """
    if status_overrides is None:
        status_overrides = {}
    
    warnings_sistema = []  # Rastreia alertas de risco trabalhista
        
    # Extrai configurações com validação segura
    jornada_minutos = settings.get('jornada_minutos', 480)
    
    # Validação: garante que jornada_minutos é um número válido
    try:
        jornada_minutos = int(jornada_minutos) if jornada_minutos else 480
        if jornada_minutos <= 0 or jornada_minutos > 1440:  # Máximo 24h
            jornada_minutos = 480
    except (ValueError, TypeError):
        jornada_minutos = 480  # Fallback para padrão
    
    tolerancia = settings.get('tolerancia', 10)
    try:
        tolerancia = int(tolerancia) if tolerancia else 10
        if tolerancia < 0 or tolerancia > 60:
            tolerancia = 10
    except (ValueError, TypeError):
        tolerancia = 10
    
    intervalo_auto = settings.get('intervalo_auto', False)
    intervalo_minutos = settings.get('intervalo_minutos', 60)
    try:
        intervalo_minutos = int(intervalo_minutos) if intervalo_minutos else 60
        if intervalo_minutos < 0 or intervalo_minutos > 480:
            intervalo_minutos = 60
    except (ValueError, TypeError):
        intervalo_minutos = 60
    sabado_util = settings.get('sabado_util', True)
    domingo_util = settings.get('domingo_util', False)
    noturno_ativo = settings.get('noturno_ativo', False)
    feriados_str = settings.get('feriados', [])
    escala_tipo = settings.get('escala_tipo', 'clt_5x2_padrao')  # Para ciclos
    data_inicio_escala = settings.get('data_inicio_escala')  # Para clt_12x36
    
    # Inicializa variáveis de feriados
    ano_detectado = datetime.now().year  # Ano atual
    feriados_set = set()  # Conjunto de feriados
    
    print(f"[DATA] Ano base: {ano_detectado}")
    
    # --- SOBRESCRITA DE META PELA ESCALA (Enterprise Fix) ---
    # Prioridade total para a definição do Catálogo, ignorando settings manuais se for padrão
    meta_sobrescrita = False
    if escala_tipo in CATALOGO_JORNADAS_CLT:
        escala_info = CATALOGO_JORNADAS_CLT[escala_tipo]
        # Se for 6x1 Comércio (440) ou 5x2 Padrão (480), força o valor correto
        if escala_tipo == 'clt_6x1_com':
            jornada_minutos = 440
            meta_sobrescrita = True
        elif escala_tipo == 'clt_5x2_padrao':
            jornada_minutos = 480
            meta_sobrescrita = True
            
        print(f"[ESCALA] Escala Ativa: {escala_info['nome']} | Meta Aplicada: {jornada_minutos}min {'(AUTO)' if meta_sobrescrita else ''}")

    if noturno_ativo:
        print(f"[NOTURNO] Adicional Noturno: ATIVO (redução Art. 73 aplicada)")

    for feriado_str in feriados_str:
        try:
            dia, mes = map(int, feriado_str.split('/'))
            feriados_set.add(date(ano_detectado, mes, dia))
        except:
            pass
            
    # Conversões de timedelta
    JORNADA_PADRAO = timedelta(minutes=jornada_minutos)
    
    # LÓGICA CICLO 12x36 v4.1: Calcula meta dinamicamente
    if escala_tipo == 'clt_12x36' and data_inicio_escala:
        try:
            # Parse data_inicio_escala (formato: DD/MM/YYYY ou YYYY-MM-DD)
            if isinstance(data_inicio_escala, str):
                if '/' in data_inicio_escala:
                    data_init = datetime.strptime(data_inicio_escala, '%d/%m/%Y').date()
                else:
                    data_init = datetime.strptime(data_inicio_escala, '%Y-%m-%d').date()
            else:
                data_init = data_inicio_escala if isinstance(data_inicio_escala, date) else data_inicio_escala.date()
            
            # Nota: A meta por dia será ajustada no loop principal para cada data_atual_obj
            ciclo_12x36_ativo = True
        except Exception as e:
            print(f"⚠️ Erro ao parsing data_inicio_escala: {e}")
            ciclo_12x36_ativo = False
    else:
        ciclo_12x36_ativo = False
    
    JORNADA_SABADO = timedelta(hours=4) if sabado_util else timedelta(0)
    TOLERANCIA = timedelta(minutes=tolerancia)
    
    df_raw = pd.DataFrame(dados_brutos)
    if df_raw.empty:
        return None, []
    
    df_raw.drop_duplicates(inplace=True)
    df_raw['data'] = pd.to_datetime(df_raw['data'])
    
    relatorio_diario = []
    todos_funcionarios = df_raw['nome'].unique()
    resumo_preview = []
    
    # NOVO v6.1: Dicionário para armazenar totais semanais por funcionário
    # Será usado pelo gerar_excel() para códigos contábeis 150/200
    totais_semanais = {}
    
    for funcionario in todos_funcionarios:
        df_funcionario_raw = df_raw[df_raw['nome'] == funcionario]
        if df_funcionario_raw.empty: continue
        
        min_date = df_funcionario_raw['data'].min().date()
        max_date = df_funcionario_raw['data'].max().date()
        periodo_completo = pd.date_range(start=min_date, end=max_date)
        
        dias_preview = []
        horas_trabalhadas_semana = {}  # Rastreador semanal para warning 44h
        
        # NOVO v6.0: Estrutura para apuração semanal de extras
        dados_semana = {}  # {num_semana: {'horas_uteis': td, 'horas_dom_fer': td, 'total': td}}
        
        for data_atual in periodo_completo:
            data_atual_obj = data_atual.date()
            dia_semana_num = data_atual_obj.weekday()
            num_semana = data_atual_obj.isocalendar()[1]
            
            # Chave para override
            override_key = f"{funcionario}|{data_atual_obj.isoformat()}"
            status_forcado = status_overrides.get(override_key)
            
            grupo = df_funcionario_raw[df_funcionario_raw['data'].dt.date == data_atual_obj]
            
            # Variáveis de cálculo
            normais = timedelta(0)
            a_dever = timedelta(0)
            extras_comuns = timedelta(0)
            extras_100 = timedelta(0)
            total_trabalhado = timedelta(0)
            adicional_noturno = timedelta(0)  # NOVO: Tracking de hora noturna
            noturno_base_minutos = 0  # NOVO v3.5: Minutos reais noturno para resumo
            
            # Batidas Separadas
            entrada_1 = None
            saida_1 = None
            entrada_2 = None
            saida_2 = None
            
            batidas_lista = []
            batidas_str = ""
            alerta = False
            status = "Normal"
            ocorrencias = ""
            
            # 1. Processamento Matemático das Batidas
            if not grupo.empty:
                horarios = [datetime.combine(data_atual_obj, h) for h in grupo["hora"]]
                horarios = ajustar_horarios_jornada_noturna(horarios, data_atual_obj, noturno_ativo=noturno_ativo)
                batidas_lista = horarios
                batidas_str = " → ".join([h.strftime("%H:%M") for h in horarios])
                
                # Validação: Batidas ímpares
                if len(horarios) % 2 != 0:
                    warning_msg = f"⚠️ {funcionario} em {data_atual_obj}: Batida ímpar ({len(horarios)} registros)"
                    warnings_sistema.append(warning_msg)
                
                # Intervalo Automático - CRÍTICO v4.6: SÓ APLICA SE EXATAMENTE 2 BATIDAS
                # NUNCA sobrescreve quando já existem 4 batidas reais!
                num_batidas_original = len(horarios)
                
                if num_batidas_original == 2 and intervalo_auto:
                    entrada, saida = horarios[0], horarios[1]
                    meio_dia = datetime.combine(data_atual_obj, dt_time(12, 0))
                    
                    # Só aplica se a jornada cruza o meio-dia
                    if entrada < meio_dia < saida:
                        fim_almoco = meio_dia + timedelta(minutes=intervalo_minutos)
                        horarios = [entrada, meio_dia, fim_almoco, saida]
                        print(f"   ⚙️ Intervalo automático aplicado: {meio_dia.strftime('%H:%M')}-{fim_almoco.strftime('%H:%M')}")
                
                # Se já tem 4+ batidas, NUNCA modifica (prioridade aos dados reais)
                elif num_batidas_original >= 4:
                    print(f"   ✅ Preservando {num_batidas_original} batidas reais (intervalo automático desabilitado)")
                
                # DISTRIBUIÇÃO DAS BATIDAS
                num_batidas = len(horarios)
                
                if num_batidas == 2:
                    entrada_1 = horarios[0].time()
                    saida_2 = horarios[1].time()
                    
                elif num_batidas == 3:
                    entrada_1 = horarios[0].time()
                    saida_1 = horarios[1].time()
                    entrada_2 = horarios[2].time()
                    ocorrencias = "BATIDA INCOMPLETA"
                    alerta = True
                    
                elif num_batidas >= 4:
                    entrada_1 = horarios[0].time()
                    saida_1 = horarios[1].time()
                    entrada_2 = horarios[2].time()
                    saida_2 = horarios[3].time()
                    
                    if num_batidas > 4:
                        ocorrencias = f"BATIDAS EXTRAS ({num_batidas})"
                
                # CÁLCULO DE HORAS REFATORADO v3.5: BASE INTEGRAL NOTURNA
                # ========================================================
                # INOVAÇÃO: A coluna "Adicional Noturno" agora exibe a BASE INTEGRAL
                # (horas reais noturnas * 1.142857) e não apenas o bônus.
                # 
                # Exemplo: 2h reais noturnas = 2 * 1.142857 = 2.285714h = 02:17:08
                
                total_segundos_clock = 0.0              # Relógio puro (sem nenhuma redução)
                base_integral_noturna_segundos = 0.0   # BASE INTEGRAL (horas * 1.142857)
                minutos_noturno_reais = 0.0            # Rastreamento de minutos reais
                
                # Processa pares sequencialmente
                for i in range(0, len(horarios) - 1, 2):
                    entrada_par = horarios[i]
                    saida_par = horarios[i + 1]
                    
                    # Calcula o tempo REAL do relógio (nunca negativo)
                    if saida_par < entrada_par:
                        saida_par = saida_par + timedelta(days=1)
                    
                    tempo_real_par = (saida_par - entrada_par).total_seconds()
                    total_segundos_clock += tempo_real_par
                    
                    # Se noturno ativo, calcula APENAS os minutos na janela [22:00-05:00]
                    if noturno_ativo:
                        # NOVO v4.2: Usa calcular_adicional_noturno_estrito para interseção precisa
                        minutos_noturno_inteiros = calcular_adicional_noturno_estrito(entrada_par, saida_par)
                        
                        if minutos_noturno_inteiros > 0:
                            minutos_noturno_reais += minutos_noturno_inteiros
                            
                            # Converte minutos reais em segundos reduzidos via fator 1.142857
                            # BASE INTEGRAL = minutos_reais * 60 * 1.142857
                            FATOR_REDUCAO_EXATO = 1.142857142857143
                            seg_reduzido = minutos_noturno_inteiros * 60 * FATOR_REDUCAO_EXATO
                            base_integral_noturna_segundos += seg_reduzido
                
                # Total Trabalhado é APENAS o relógio real (PROIBIDO somar ganho noturno)
                total_trabalhado = timedelta(seconds=total_segundos_clock)
                
                # HOTFIX v6.0: Adicional Noturno é o tempo REAL (relógio), NÃO reduzido
                # O sistema de folha do contador é que aplica o fator 1.1428
                # Isso bate com o Fechamento Exemplar que mostra 2:03 em vez de 2:19
                if noturno_ativo and minutos_noturno_reais > 0:
                    adicional_noturno = timedelta(minutes=minutos_noturno_reais)  # REAL
                else:
                    adicional_noturno = timedelta(0)
                
                # Rastreamento para o frontend (minutos reais noturno para cálculo de resumo)
                noturno_base_minutos = minutos_noturno_reais
                
                # INICIALIZAÇÃO CRÍTICA v4.0: meta_dia deve estar disponível em TODOS os caminhos
                meta_dia = calcular_meta_dinamica_escala(
                    escala_tipo, 
                    dia_semana_num, 
                    data_atual_obj,
                    date.fromisoformat(data_init) if ciclo_12x36_ativo and isinstance(data_init, date) else None
                )
                
                # Validação: Intervalo intrajornada < 1h (Art. 71 CLT)
                if len(horarios) >= 4:
                    intervalo = horarios[2] - horarios[1]
                    if intervalo < timedelta(hours=1):
                        warning_msg = f"⚠️ {funcionario} em {data_atual_obj}: Intervalo < 1h ({intervalo.total_seconds()/60:.0f}min) - Risco Art. 71"
                        warnings_sistema.append(warning_msg)
                
                # Classificação Automática
                eh_feriado = data_atual_obj in feriados_set
                eh_domingo = dia_semana_num == 6
                eh_sabado = dia_semana_num == 5
                
                # CRÍTICO v4.7: Ajusta meta_dia para ZERO em dias não úteis
                # Isso evita que a fórmula Excel calcule faltas em domingos/feriados
                if eh_feriado or (eh_domingo and not domingo_util):
                    meta_dia = timedelta(0)  # ZERO meta em feriados/domingos não úteis
                    extras_100 = total_trabalhado
                    status = "Extra 100%" if total_trabalhado > timedelta(0) else "Feriado"
                elif eh_sabado:
                    # HOTFIX v6.0: Sábado USA a meta da escala (7h20 para 6x1), NÃO 4h fixo
                    # meta_dia já foi calculada corretamente por calcular_meta_dinamica_escala
                    # Apenas zera se sabado_util=False E escala 5x2
                    if not sabado_util and escala_tipo == 'clt_5x2_padrao':
                        meta_dia = timedelta(0)
                        extras_100 = total_trabalhado
                        status = "Extra 100%" if total_trabalhado > timedelta(0) else "Folga"
                    else:
                        # Sábado útil: usa meta_dia da escala (ex: 440min para 6x1)
                        normais = min(total_trabalhado, meta_dia)
                        if total_trabalhado > meta_dia:
                            extras_comuns = total_trabalhado - meta_dia
                            status = "Extra"
                        else:
                            a_dever = meta_dia - total_trabalhado
                            status = "Incompleto" if a_dever > timedelta(0) else "Normal"
                else:
                    # Ajusta meta dinamicamente baseada no tipo de escala (já inicializado acima)
                    # meta_dia já foi calculada na inicialização crítica v4.0
                    
                    normais = min(total_trabalhado, meta_dia)
                    if total_trabalhado > meta_dia:
                        extras_comuns = total_trabalhado - meta_dia
                        status = "Extra"
                    else:
                        a_dever = meta_dia - total_trabalhado
                        status = "Incompleto" if a_dever > TOLERANCIA else "Normal"
                        alerta = a_dever > TOLERANCIA
                    
                    # INTEGRAÇÃO v4.1: Aplica tolerância Art. 58 §1º (VTD)
                    vtd_minutos = (total_trabalhado.total_seconds() - meta_dia.total_seconds()) / 60.0
                    minutos_abonados, minutos_descontados, obs_vtd = aplicar_tolerancia_clt(vtd_minutos)
                    
                    if minutos_descontados > 0:
                        # VTD positivo e > 10min: desconta integral
                        status = "Incompleto"
                        alerta = True
                    elif minutos_abonados > 0:
                        # VTD negativo e < -10min: abona integral
                        status = "Extra"
                        alerta = False
                    elif abs(vtd_minutos) <= 10:
                        # VTD dentro da tolerância: normal
                        status = "Normal" if status != "Incompleto" else "Incompleto"
                        alerta = False

            else:
                # Sem batidas - INICIALIZA meta_dia para garantir que está definida
                meta_dia = calcular_meta_dinamica_escala(
                    escala_tipo, 
                    dia_semana_num, 
                    data_atual_obj,
                    date.fromisoformat(data_init) if ciclo_12x36_ativo and isinstance(data_init, date) else None
                )
                
                if data_atual_obj in feriados_set or dia_semana_num == 6:
                    status = "Folga"
                    meta_dia = timedelta(0)  # Zera meta em feriados/domingos
                    ocorrencias = "DSR/FERIADO"
                elif dia_semana_num == 5 and not sabado_util:
                    status = "Folga"
                    meta_dia = timedelta(0)
                    ocorrencias = "DSR"
                else:
                    status = "Falta"
                    a_dever = JORNADA_SABADO if (dia_semana_num == 5) else JORNADA_PADRAO
                    alerta = True
                    batidas_str = "Falta"
                    ocorrencias = "FALTA NÃO JUSTIFICADA"

            # 2. APLICAÇÃO DE STATUS MANUAL
            if status_forcado:
                status = status_forcado
                
                if status == 'ABONO':
                    print(f" ��️ Aplicando ABONO em {data_atual_obj}: Zerando dívida de {a_dever}")
                    a_dever = timedelta(0)
                    alerta = False
                    ocorrencias = "ABONADO"
                    
                elif status == 'ATESTADO':
                    a_dever = timedelta(0)
                    alerta = False
                    ocorrencias = "ATESTADO MÉDICO"
                    if not batidas_lista:
                        batidas_str = "Atestado"
                
                elif status in ['FOLGA', 'FERIADO', 'DSR']:
                    a_dever = timedelta(0)
                    alerta = False
                    ocorrencias = status.upper()
                    if not batidas_lista:
                        batidas_str = status.title()
                
                elif status == 'FALTA':
                    if not batidas_lista:
                        a_dever = JORNADA_SABADO if (dia_semana_num == 5) else JORNADA_PADRAO
                    alerta = True
                    ocorrencias = "FALTA"

            # Calcula saldo final do dia
            saldo_dia = extras_comuns + extras_100 - a_dever
            
            # Rastreador semanal
            if num_semana not in horas_trabalhadas_semana:
                horas_trabalhadas_semana[num_semana] = timedelta(0)
            horas_trabalhadas_semana[num_semana] += total_trabalhado
            
            # NOVO v6.0: Coletar dados para apuração semanal de extras
            if num_semana not in dados_semana:
                dados_semana[num_semana] = {
                    'horas_uteis': timedelta(0),
                    'horas_dom_fer': timedelta(0),
                    'total': timedelta(0)
                }
            
            # Classificar horas: domingo/feriado vs dias úteis
            eh_dom_fer_para_semana = (data_atual_obj in feriados_set) or (dia_semana_num == 6 and not domingo_util)
            if eh_dom_fer_para_semana:
                dados_semana[num_semana]['horas_dom_fer'] += total_trabalhado
            else:
                dados_semana[num_semana]['horas_uteis'] += total_trabalhado
            dados_semana[num_semana]['total'] += total_trabalhado
            
            # Monta registro para Excel (ESTRUTURA ATUALIZADA)
            relatorio_diario.append({
                "Data": data_atual_obj,
                "Funcionário": funcionario,
                "Dia da Semana": DIAS_SEMANA.get(dia_semana_num, ''),
                "Entrada 1": entrada_1,
                "Saída 1": saida_1,
                "Entrada 2": entrada_2,
                "Saída 2": saida_2,
                "Meta": meta_dia,  # NOVO v4.0: Coluna Meta para cada dia
                "Total Trabalhado": total_trabalhado,
                "Adicional Noturno": adicional_noturno,  # NOVO
                "Horas Normais": normais,
                "Horas a Dever": a_dever,
                "Horas Extras (Comum)": extras_comuns,
                "Horas Extras (100%)": extras_100,
                "Ocorrências": ocorrencias
            })
            
            # Preview (mantém string para o frontend)
            # ATUALIZADO v4.0: Adicionado tipo_dia, meta_minutos, banco_horas_informativo
            tipo_dia_str = "normal"
            if data_atual_obj in feriados_set:
                tipo_dia_str = "feriado"
            elif dia_semana_num == 6:  # Domingo
                tipo_dia_str = "descanso"
            elif dia_semana_num == 5:  # Sábado
                tipo_dia_str = "sabado"
            
            dias_preview.append({
                "data": data_atual_obj.strftime("%d/%m"),
                "dia_semana": DIAS_SEMANA.get(dia_semana_num, '')[:3],
                "batidas": batidas_str,
                "total": format_td(total_trabalhado),
                "noturno_base": noturno_base_minutos,
                "batidas_4cols": {
                    "entrada_1": entrada_1.strftime("%H:%M") if entrada_1 else None,
                    "saida_1": saida_1.strftime("%H:%M") if saida_1 else None,
                    "entrada_2": entrada_2.strftime("%H:%M") if entrada_2 else None,
                    "saida_2": saida_2.strftime("%H:%M") if saida_2 else None
                },
                "saldo": format_td(saldo_dia),
                "status": status,
                "alerta": alerta,
                # NOVOS CAMPOS v4.0
                "tipo_dia": tipo_dia_str,
                "meta_minutos": int(meta_dia.total_seconds() / 60),
                "banco_horas_informativo": True  # Flag de segurança
            })
        
        # Validação: Semanas com > 44h trabalhadas
        for num_semana, horas in horas_trabalhadas_semana.items():
            if horas > timedelta(hours=44):
                warning_msg = f"⚠️ {funcionario} (semana {num_semana}): {format_td(horas)} > 44h - Risco trabalhista"
                warnings_sistema.append(warning_msg)
        
        # Totais - NOVO v6.0: Apuração Semanal de Extras
        df_func = pd.DataFrame([r for r in relatorio_diario if r["Funcionário"] == funcionario])
        
        # Lê configuração de regra de cálculo (default: semanal)
        regra_extra = settings.get('regra_extra', 'semanal')
        jornada_semanal_minutos = settings.get('jornada_semanal_minutos', 2640)  # 44h default
        
        if regra_extra == 'semanal' and dados_semana:
            # NOVO v6.0: Apuração semanal CLT (44h com compensação)
            debug_mode = settings.get('debug_calculo', True)  # Default True para rollout
            
            if debug_mode:
                print(f"\n📊 APURAÇÃO SEMANAL - {funcionario}")
            
            extras_50_total, extras_100_total = calcular_extras_semanal(
                dados_semana, 
                jornada_semanal_minutos,
                debug=debug_mode
            )
            
            totals = {
                "Normais": df_func["Horas Normais"].sum() if not df_func.empty else timedelta(),
                "A Dever": df_func["Horas a Dever"].sum() if not df_func.empty else timedelta(),
                "Extras Comum": extras_50_total,
                "Extras 100%": extras_100_total,
                "Noturno": df_func["Adicional Noturno"].sum() if not df_func.empty else timedelta(),
            }
            
            # NOVO v6.1: Armazena totais semanais para uso no Excel
            # IMPORTANTE: Estes valores já vêm da apuração semanal CLT (44h),
            # a mesma usada no preview JSON. NÃO recalcular no Excel.
            totais_semanais[funcionario] = {
                "extra50": extras_50_total,
                "extra100": extras_100_total
            }
            
            # Log de resultado (controlado por debug_mode)
            if debug_mode:
                e50_h = extras_50_total.total_seconds() / 3600
                e100_h = extras_100_total.total_seconds() / 3600
                print(f"  ✅ RESULTADO: Extra 50% = {e50_h:.2f}h | Extra 100% = {e100_h:.2f}h | Total Extras = {e50_h + e100_h:.2f}h")
        else:
            # Modo diário (legado)
            totals = {
                "Normais": df_func["Horas Normais"].sum() if not df_func.empty else timedelta(),
                "A Dever": df_func["Horas a Dever"].sum() if not df_func.empty else timedelta(),
                "Extras Comum": df_func["Horas Extras (Comum)"].sum() if not df_func.empty else timedelta(),
                "Extras 100%": df_func["Horas Extras (100%)"].sum() if not df_func.empty else timedelta(),
                "Noturno": df_func["Adicional Noturno"].sum() if not df_func.empty else timedelta(),
            }
        
        saldo_final = totals["Extras Comum"] + totals["Extras 100%"] - totals["A Dever"]
        
        resumo_preview.append({
            "funcionario": funcionario,
            "normais": format_td(totals["Normais"]).replace("+", ""),
            "dever": format_td(totals["A Dever"]).replace("+", ""),
            "extras_comuns": format_td(totals["Extras Comum"]).replace("+", ""),
            "extras_100": format_td(totals["Extras 100%"]).replace("+", ""),
            "saldo": format_td(saldo_final),
            "dias": dias_preview,
            # CAMPOS v6.0 - Apuração Semanal CLT
            "aviso_saldo": "Extras calculados com apuração semanal de 44h.",
            "saldo_eh_informativo": True,
            "versao_calculo": "v6.0-semanal-clt",
            "regra_extra": regra_extra
        })
        
    return relatorio_diario, resumo_preview, totais_semanais

# ===== FUNÇÃO REFATORADA: GERAR EXCEL PROFISSIONAL =====
def gerar_excel(relatorio_diario: List[dict], settings: dict = None, totais_semanais: dict = None) -> io.BytesIO:
    """
    Gera arquivo Excel profissional estilo "Espelho de Ponto" do Departamento Pessoal.
    
    MUDANÇAS CRÍTICAS:
    - Batidas separadas em 4 colunas (Ent.1, Sai.1, Ent.2, Sai.2)
    - Valores numéricos reais (float) com formatação [h]:mm:ss
    - Fórmulas SUM no rodapé
    - Design profissional (cabeçalho, bordas, zebrado)
    - Hash SHA-256 de integridade
    
    NOVO v6.1: Códigos contábeis 150/200 usam totais semanais do backend
    (mesma apuração CLT 44h do preview JSON)
    """
    if settings is None:
        settings = {}
    if totais_semanais is None:
        totais_semanais = {}
    
    df_calculado = pd.DataFrame(relatorio_diario)
    
    output = io.BytesIO()
    wb = Workbook()
    
    # Remove planilha padrão
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    if not df_calculado.empty:
        df_calculado.sort_values(by=['Funcionário', 'Data'], inplace=True)
        
        for funcionario in df_calculado["Funcionário"].unique():
            df_funcionario = df_calculado[df_calculado["Funcionário"] == funcionario].copy()
            
            # Cria planilha
            ws = wb.create_sheet(title=funcionario[:31])
            
            # --- CABEÇALHO PROFISSIONAL ---
            empresa_nome = settings.get('empresa_nome', 'EMPRESA LTDA')
            empresa_cnpj = settings.get('empresa_cnpj', '00.000.000/0000-00')
            
            # Linha 1: Nome da Empresa (Fundo Escuro)
            ws.merge_cells('A1:M1')
            cell_empresa = ws['A1']
            cell_empresa.value = empresa_nome.upper()
            cell_empresa.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
            cell_empresa.fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
            cell_empresa.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Linha 2: CNPJ
            ws.merge_cells('A2:M2')
            cell_cnpj = ws['A2']
            cell_cnpj.value = f"CNPJ: {empresa_cnpj}"
            cell_cnpj.font = Font(name='Arial', size=10, color='FFFFFF')
            cell_cnpj.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            cell_cnpj.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 20
            
            # Linha 3: Título Espelho de Ponto
            ws.merge_cells('A3:M3')
            cell_titulo = ws['A3']
            cell_titulo.value = "ESPELHO DE PONTO - REGISTRO DE HORÁRIOS"
            cell_titulo.font = Font(name='Arial', size=12, bold=True)
            cell_titulo.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[3].height = 22
            
            # Linha 4: Período e Funcionário
            min_date = df_funcionario['Data'].min()
            max_date = df_funcionario['Data'].max()
            ws.merge_cells('A4:M4')
            cell_periodo = ws['A4']
            cell_periodo.value = f"Período: {min_date.strftime('%d/%m/%Y')} a {max_date.strftime('%d/%m/%Y')} | Funcionário: {funcionario.upper()}"
            cell_periodo.font = Font(name='Arial', size=10, italic=True)
            cell_periodo.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[4].height = 18
            
            # Linha 5: Espaço
            ws.row_dimensions[5].height = 10
            
            # --- CABEÇALHOS DA TABELA (Linha 6) ---
            headers = ['Data', 'Dia', 'Ent. 1', 'Sai. 1', 'Ent. 2', 'Sai. 2', 
                      'Meta', 'Total', 'Noturno', 'Normais', 'Faltas', 'Extra 50%', 'Extra 100%', 'Ocorrências']
            
            thin_border = Border(
                left=Side(style='thin', color='BDC3C7'),
                right=Side(style='thin', color='BDC3C7'),
                top=Side(style='thin', color='BDC3C7'),
                bottom=Side(style='thin', color='BDC3C7')
            )
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=6, column=col_idx, value=header)
                cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='1ABC9C', end_color='1ABC9C', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            ws.row_dimensions[6].height = 20
            
            # --- DADOS (A partir da linha 7) ---
            start_row = 7
            
            for idx, row_data in df_funcionario.iterrows():
                row_num = start_row + (idx - df_funcionario.index[0])
                
                # Coluna A: Data (Texto)
                ws.cell(row=row_num, column=1, value=row_data['Data'].strftime('%d/%m/%Y'))
                ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
                ws.cell(row=row_num, column=1).border = thin_border
                
                # Coluna B: Dia da Semana (Texto)
                dia_semana_abrev = row_data['Dia da Semana'][:3].upper()
                cell_dia = ws.cell(row=row_num, column=2, value=dia_semana_abrev)
                cell_dia.alignment = Alignment(horizontal='center')
                cell_dia.border = thin_border
                
                # Destaque para finais de semana
                if dia_semana_abrev in ['SÁB', 'DOM']:
                    cell_dia.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
                
                # Colunas C-F: Batidas (Valores Numéricos de Tempo)
                batidas_cols = [
                    ('Entrada 1', 3),
                    ('Saída 1', 4),
                    ('Entrada 2', 5),
                    ('Saída 2', 6)
                ]
                
                for col_name, col_idx in batidas_cols:
                    valor_time = row_data.get(col_name)
                    cell = ws.cell(row=row_num, column=col_idx)
                    
                    if valor_time is not None:
                        # Converte datetime.time para fração de dia do Excel
                        cell.value = time_to_excel_time(valor_time)
                        cell.number_format = 'HH:MM'
                    else:
                        cell.value = None
                    
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                # Coluna G: Meta Diária (NOVO v4.0)
                cell_meta = ws.cell(row=row_num, column=7)
                valor_meta = row_data.get('Meta', timedelta(0))
                cell_meta.value = timedelta_to_excel_time(valor_meta)
                cell_meta.number_format = '[h]:mm:ss'
                cell_meta.alignment = Alignment(horizontal='center')
                cell_meta.border = thin_border
                
                # Coluna H: Total Trabalhado - FÓRMULA DINÂMICA RESILIENTE (v4.8)
                cell_total = ws.cell(row=row_num, column=8)
                
                # NOVA FÓRMULA v4.8: Usa COUNT() para evitar #VALUE! quando células vazias
                # Formato: IF(COUNT(range)<2, 0, cálculo_normal)
                # Se contador apagar uma batida, mostra 0 em vez de erro
                
                # Primeiro turno: C=Entrada1, D=Saída1
                # Segundo turno: E=Entrada2, F=Saída2
                formula_turno1 = f"IF(COUNT(C{row_num}:D{row_num})<2,0,IF(D{row_num}<C{row_num},D{row_num}+1-C{row_num},D{row_num}-C{row_num}))"
                formula_turno2 = f"IF(COUNT(E{row_num}:F{row_num})<2,0,IF(F{row_num}<E{row_num},F{row_num}+1-E{row_num},F{row_num}-E{row_num}))"
                
                # Combina os dois turnos
                cell_total.value = f"={formula_turno1}+{formula_turno2}"
                
                cell_total.number_format = '[h]:mm:ss'
                cell_total.alignment = Alignment(horizontal='center')
                cell_total.border = thin_border
                
                # Coluna I: Adicional Noturno (Deslocado de H para I - Art. 73 CLT)
                cell_noturno = ws.cell(row=row_num, column=9)
                valor_noturno = row_data.get('Adicional Noturno', timedelta(0))
                cell_noturno.value = timedelta_to_excel_time(valor_noturno)
                cell_noturno.number_format = '[h]:mm:ss'
                cell_noturno.alignment = Alignment(horizontal='center')
                cell_noturno.border = thin_border
                # Destaque visual para hora noturna
                if valor_noturno > timedelta(0):
                    cell_noturno.fill = PatternFill(start_color='FEF5E7', end_color='FEF5E7', fill_type='solid')
                    cell_noturno.font = Font(bold=True, color='D68910')
                
                # Colunas J-M: Horas com FÓRMULAS DINÂMICAS (CRÍTICO v4.6)
                # J = Horas Normais: MIN(Total, Meta)
                # K = Horas a Dever (Faltas): MAX(0, Meta - Total)
                # L = Horas Extras 50%: MAX(0, Total - Meta)
                # M = Horas Extras 100%: Calculado separadamente (DSR, feriados)
                
                # Coluna J: Horas Normais - FÓRMULA DINÂMICA
                cell_normais = ws.cell(row=row_num, column=10)
                cell_normais.value = f"=MIN(H{row_num},G{row_num})"  # MIN(Total, Meta)
                cell_normais.number_format = '[h]:mm:ss'
                cell_normais.alignment = Alignment(horizontal='center')
                cell_normais.border = thin_border
                
                # Coluna K: Horas a Dever (Faltas) - FÓRMULA DINÂMICA
                cell_faltas = ws.cell(row=row_num, column=11)
                cell_faltas.value = f"=MAX(0,G{row_num}-H{row_num})"  # MAX(0, Meta - Total)
                cell_faltas.number_format = '[h]:mm:ss'
                cell_faltas.alignment = Alignment(horizontal='center')
                cell_faltas.border = thin_border
                cell_faltas.font = Font(color='C0392B', bold=True)  # Vermelho para faltas
                
                # Coluna L: Horas Extras 50% - FÓRMULA DINÂMICA
                cell_extra50 = ws.cell(row=row_num, column=12)
                cell_extra50.value = f"=MAX(0,H{row_num}-G{row_num})"  # MAX(0, Total - Meta)
                cell_extra50.number_format = '[h]:mm:ss'
                cell_extra50.alignment = Alignment(horizontal='center')
                cell_extra50.border = thin_border
                
                # Coluna M: Horas Extras 100% - VALOR ESTÁTICO (cálculo complexo de DSR/feriados)
                cell_extra100 = ws.cell(row=row_num, column=13)
                valor_extra100 = row_data.get('Horas Extras (100%)', timedelta(0))
                cell_extra100.value = timedelta_to_excel_time(valor_extra100)
                cell_extra100.number_format = '[h]:mm:ss'
                cell_extra100.alignment = Alignment(horizontal='center')
                cell_extra100.border = thin_border
                
                # Coluna N: Ocorrências (Texto) - DESLOCADA PARA 14
                cell_ocorr = ws.cell(row=row_num, column=14, value=row_data.get('Ocorrências', ''))
                cell_ocorr.alignment = Alignment(horizontal='left', wrap_text=True)
                cell_ocorr.font = Font(size=8)
                cell_ocorr.border = thin_border
                
                # Zebrado (linhas alternadas)
                if (row_num - start_row) % 2 == 1:
                    for col in range(1, 15):  # Aumentado para 15 (incluindo coluna Meta)
                        ws.cell(row=row_num, column=col).fill = PatternFill(
                            start_color='F8F9FA', end_color='F8F9FA', fill_type='solid'
                        )
            
            last_data_row = start_row + len(df_funcionario) - 1
            
            # --- RODAPÉ COM TOTAIS (Fórmulas) ---
            total_row = last_data_row + 2
            
            ws.cell(row=total_row, column=1, value="TOTAIS:")
            ws.cell(row=total_row, column=1).font = Font(bold=True, size=11)
            ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right')
            
            # Fórmulas SUM para cada coluna de tempo - DESLOCADAS UMA COLUNA (incluindo Meta)
            formulas_cols = [
                (7, 'G'),   # Meta
                (8, 'H'),   # Total
                (9, 'I'),   # Noturno
                (10, 'J'),   # Normais
                (11, 'K'),  # Faltas
                (12, 'L'),  # Extra 50%
                (13, 'M')   # Extra 100%
            ]
            
            for col_idx, col_letter in formulas_cols:
                cell = ws.cell(row=total_row, column=col_idx)
                cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
                cell.number_format = '[h]:mm:ss'
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                cell.fill = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid')
            
            # Saldo Final (Extra Comum + Extra 100% - Faltas)
            saldo_row = total_row + 1
            ws.merge_cells(f'A{saldo_row}:B{saldo_row}')
            cell_saldo_label = ws.cell(row=saldo_row, column=1, value="SALDO HORAS EXCEDENTES (Informativo 1:1):")
            cell_saldo_label.font = Font(bold=True, size=11, color='1ABC9C')
            cell_saldo_label.alignment = Alignment(horizontal='right')
            
            ws.merge_cells(f'L{saldo_row}:M{saldo_row}')
            cell_saldo_valor = ws.cell(row=saldo_row, column=12)  # DESLOCADO: era 11, agora 12
            cell_saldo_valor.value = f"=L{total_row}+M{total_row}-K{total_row}"  # Extra 50% + 100% - Faltas
            cell_saldo_valor.number_format = '[h]:mm:ss'
            cell_saldo_valor.font = Font(bold=True, size=13, color='1ABC9C')
            cell_saldo_valor.alignment = Alignment(horizontal='center')
            cell_saldo_valor.border = Border(
                top=Side(style='thick', color='1ABC9C'),
                bottom=Side(style='thick', color='1ABC9C'),
                left=Side(style='thick', color='1ABC9C'),
                right=Side(style='thick', color='1ABC9C')
            )
            
            # --- CÓDIGOS CONTÁBEIS (PADRÃO FECHAMENTO DE FOLHA) ---
            codigos_row = saldo_row + 2
            
            # Cabeçalho da seção
            ws.merge_cells(f'A{codigos_row}:N{codigos_row}')
            cell_cod_header = ws.cell(row=codigos_row, column=1, value="RESUMO PARA FECHAMENTO DE FOLHA (CÓDIGOS CONTÁBEIS)")
            cell_cod_header.font = Font(bold=True, size=11, color='FFFFFF')
            cell_cod_header.fill = PatternFill(start_color='34495E', end_color='34495E', fill_type='solid')
            cell_cod_header.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[codigos_row].height = 22
            
            # Linha de cabeçalhos das colunas
            cod_header_row = codigos_row + 1
            headers_cod = ['Código', 'Descrição', 'Horas (HH:MM)', 'Horas Decimais']
            for col_idx, header in enumerate(headers_cod, start=1):
                cell = ws.cell(row=cod_header_row, column=col_idx, value=header)
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # IMPORTANTE v6.1:
            # Os valores abaixo já vêm da apuração semanal CLT (44h),
            # a mesma usada no preview JSON.
            # NÃO usar fórmulas Excel que referenciam soma diária para evitar divergências.
            
            # Obtém totais semanais do backend (ou fallback para timedelta zero)
            func_totais = totais_semanais.get(funcionario, {})
            extras_50_semanal = func_totais.get("extra50", timedelta())
            extras_100_semanal = func_totais.get("extra100", timedelta())
            
            # Código 150 - Extra 50% (Dias Úteis) - VALOR SEMANAL CLT
            cod150_row = cod_header_row + 1
            ws.cell(row=cod150_row, column=1, value="150").alignment = Alignment(horizontal='center')
            ws.cell(row=cod150_row, column=2, value="Horas Extras 50% (Dias Úteis)")
            cell_150_horas = ws.cell(row=cod150_row, column=3)
            cell_150_horas.value = timedelta_to_excel_time(extras_50_semanal)  # Valor semanal CLT
            cell_150_horas.number_format = '[h]:mm:ss'
            cell_150_horas.alignment = Alignment(horizontal='center')
            cell_150_decimal = ws.cell(row=cod150_row, column=4)
            cell_150_decimal.value = extras_50_semanal.total_seconds() / 3600  # Decimal semanal
            cell_150_decimal.number_format = '0.00'
            cell_150_decimal.alignment = Alignment(horizontal='center')
            
            # Código 200 - Extra 100% (Domingos/Feriados) - VALOR SEMANAL CLT
            cod200_row = cod150_row + 1
            ws.cell(row=cod200_row, column=1, value="200").alignment = Alignment(horizontal='center')
            ws.cell(row=cod200_row, column=2, value="Horas Extras 100% (Domingos/Feriados)")
            cell_200_horas = ws.cell(row=cod200_row, column=3)
            cell_200_horas.value = timedelta_to_excel_time(extras_100_semanal)  # Valor semanal CLT
            cell_200_horas.number_format = '[h]:mm:ss'
            cell_200_horas.alignment = Alignment(horizontal='center')
            cell_200_decimal = ws.cell(row=cod200_row, column=4)
            cell_200_decimal.value = extras_100_semanal.total_seconds() / 3600  # Decimal semanal
            cell_200_decimal.number_format = '0.00'
            cell_200_decimal.alignment = Alignment(horizontal='center')
            
            # Código 25 - Adicional Noturno — SEÇÃO SEPARADA (INFORMATIVO)
            # ============================================================
            # CORREÇÃO v6.0: Noturno é ATRIBUTO, não TEMPO somável.
            # Separar visualmente para evitar que usuário some incorretamente.
            # ============================================================
            
            # Linha de separação visual
            noturno_header_row = cod200_row + 2
            ws.merge_cells(f'A{noturno_header_row}:D{noturno_header_row}')
            cell_noturno_header = ws.cell(row=noturno_header_row, column=1)
            cell_noturno_header.value = "INFORMATIVO — ADICIONAL NOTURNO (não soma em horas)"
            cell_noturno_header.font = Font(bold=True, size=9, italic=True, color='7F8C8D')
            cell_noturno_header.fill = PatternFill(start_color='ECF0F1', end_color='ECF0F1', fill_type='solid')
            cell_noturno_header.alignment = Alignment(horizontal='center', vertical='center')
            cell_noturno_header.border = thin_border
            ws.row_dimensions[noturno_header_row].height = 18
            
            # Código 25 - Dados
            cod25_row = noturno_header_row + 1
            ws.cell(row=cod25_row, column=1, value="25").alignment = Alignment(horizontal='center')
            ws.cell(row=cod25_row, column=2, value="Base Noturna (atributo financeiro)")
            cell_25_horas = ws.cell(row=cod25_row, column=3)
            cell_25_horas.value = f"=I{total_row}"  # Referência ao total de Noturno
            cell_25_horas.number_format = '[h]:mm:ss'
            cell_25_horas.alignment = Alignment(horizontal='center')
            cell_25_decimal = ws.cell(row=cod25_row, column=4)
            cell_25_decimal.value = f"=I{total_row}*24"  # Noturno em decimal
            cell_25_decimal.number_format = '0.00'
            cell_25_decimal.alignment = Alignment(horizontal='center')
            
            # Aplicar bordas em todas as células dos códigos (150 e 200 apenas)
            for row in range(cod_header_row, cod200_row + 1):
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = thin_border
            
            # Bordas para seção noturno
            for col in range(1, 5):
                ws.cell(row=cod25_row, column=col).border = thin_border
            
            # Destaque visual para códigos de TEMPO (150 e 200)
            for row in [cod150_row, cod200_row]:
                ws.cell(row=row, column=1).font = Font(bold=True, color='2C3E50')
                ws.cell(row=row, column=4).font = Font(bold=True, color='27AE60')
            
            # Destaque visual diferenciado para noturno (cor mais suave)
            ws.cell(row=cod25_row, column=1).font = Font(bold=True, color='7F8C8D')
            ws.cell(row=cod25_row, column=4).font = Font(bold=True, color='D68910')
            
            # --- AVISO LEGAL (SIMPLIFICADO) ---
            aviso_row = cod25_row + 2
            ws.merge_cells(f'A{aviso_row}:N{aviso_row}')
            cell_aviso = ws.cell(row=aviso_row, column=1)
            cell_aviso.value = (
                "Códigos 150 e 200 = horas para pagamento. "
                "Código 25 = base para adicional noturno (não soma em horas trabalhadas). "
                "Saldo 1:1 é informativo."
            )
            cell_aviso.font = Font(size=8, italic=True, color='7F8C8D')
            cell_aviso.alignment = Alignment(wrap_text=True, horizontal='left')
            ws.row_dimensions[aviso_row].height = 20
            
            # --- HASH DE INTEGRIDADE ---
            hash_row = aviso_row + 2
            
            # Gera hash SHA-256 dos dados
            dados_str = json.dumps(df_funcionario.to_dict(), sort_keys=True, default=str)
            hash_value = hashlib.sha256(dados_str.encode()).hexdigest()[:16]
            
            ws.merge_cells(f'A{hash_row}:N{hash_row}')  # DESLOCADO: era M, agora N
            cell_hash = ws.cell(row=hash_row, column=1)
            cell_hash.value = f"�� Hash de Integridade: {hash_value.upper()}"
            cell_hash.font = Font(size=7, color='95A5A6', name='Courier New')
            cell_hash.alignment = Alignment(horizontal='center')
            
            # --- ASSINATURAS ---
            assinatura_row = hash_row + 3
            
            ws.merge_cells(f'A{assinatura_row}:E{assinatura_row}')
            ws.cell(row=assinatura_row, column=1, value="_" * 40)
            ws.cell(row=assinatura_row + 1, column=1, value="ASSINATURA DO FUNCIONÁRIO")
            ws.cell(row=assinatura_row + 1, column=1).font = Font(size=9, bold=True)
            ws.cell(row=assinatura_row + 1, column=1).alignment = Alignment(horizontal='center')
            ws.merge_cells(f'H{assinatura_row}:L{assinatura_row}')
        ws.cell(row=assinatura_row, column=8, value="_" * 40)
        ws.cell(row=assinatura_row + 1, column=8, value="ASSINATURA DO GESTOR")
        ws.cell(row=assinatura_row + 1, column=8).font = Font(size=9, bold=True)
        ws.cell(row=assinatura_row + 1, column=8).alignment = Alignment(horizontal='center')
        
        # --- AJUSTE DE LARGURAS DAS COLUNAS ---
        column_widths = {
            'A': 12,  # Data
            'B': 8,   # Dia
            'C': 9,   # Ent. 1
            'D': 9,   # Sai. 1
            'E': 9,   # Ent. 2
            'F': 9,   # Sai. 2
            'G': 10,  # Meta (NOVO)
            'H': 10,  # Total
            'I': 10,  # Noturno
            'J': 10,  # Normais
            'K': 10,  # Faltas
            'L': 11,  # Extra 50%
            'M': 11,  # Extra 100%
            'N': 25   # Ocorrências
        }
        
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Configurações de impressão (Paisagem, A4)
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        
        # Margens
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75
        
        # Congela painéis (cabeçalho)
        ws.freeze_panes = 'A7'
    
    wb.save(output)
    output.seek(0)
    return output


# ===== ROTAS DA API =====
@app.get("/")
async def root():
    """Health check endpoint"""
    num_keys = len(GEMINI_MODELS)
    if num_keys == 0:
        gemini_status = "⚠️ Não configurado"
    elif num_keys == 1:
        gemini_status = "✅ 1 chave configurada"
    else:
        gemini_status = f"✅ {num_keys} chaves configuradas (fallback automático)"
    
    return {
        "status": "online",
        "versao": "2.2 - Múltiplas Chaves API",
        "gemini_api": gemini_status,
        "modelo": "gemini-2.5-flash",
        "formatos_suportados": ["TXT", "PDF", "JPG", "PNG"],
        "recursos": {
            "jornada_noturna": "✅ Suportada",
            "feriados_dinamicos": "✅ Frontend-driven",
            "edicao_manual": "✅ Rota /recalcular",
            "intervalo_auto": "✅ Suportado",
            "excel_profissional": "✅ Espelho de Ponto Corporativo",
            "batidas_separadas": "✅ 4 colunas editáveis",
            "formulas_excel": "✅ Totais dinâmicos",
            "status_opcoes": "FALTA, ATESTADO, FOLGA, DSR, FERIADO, ABONO",
            "multi_api_keys": f"✅ Suporte a {num_keys} chave(s) com fallback automático"
        },
        "status_explicacoes": {
            "FALTA": "Dia não trabalhado sem justificativa",
            "ATESTADO": "Dia abonado por atestado médico",
            "FOLGA": "Dia de folga programada",
            "DSR": "Descanso semanal remunerado",
            "FERIADO": "Dia feriado",
            "ABONO": "Perdoa atraso/saldo negativo mantendo batidas"
        }
    }


@app.post("/converter")
async def converter_cartao_ponto(
    files: List[UploadFile] = File(...),
    settings: str = Form(...),
    consent_metadata: str = Form(...)
):
    """
    Endpoint principal: processa arquivos, aplica settings e retorna preview + Excel
    """
    try:
        # Parse dos settings
        settings_dict = json.loads(settings)
        consent_dict = json.loads(consent_metadata)
        
        print(f"\n{'='*70}")
        print(f"[IN] Processando {len(files)} arquivo(s)")
        print(f"[CFG] Settings: Jornada {settings_dict.get('jornada_minutos')}min, Feriados: {len(settings_dict.get('feriados', []))}")
        print(f"[OK] Consentimento LGPD registrado: {consent_dict.get('timestamp')}")
        print(f"{'='*70}\n")
        
        # Processa todos os arquivos
        dados_consolidados = []
        
        for arquivo in files:
            filename = arquivo.filename.lower()
            
            try:
                if filename.endswith('.txt'):
                    conteudo = arquivo.file.read().decode("utf-8")
                    arquivo.file.seek(0)
                    dados = processar_txt(conteudo)
                    dados_consolidados.extend(dados)
                    
                elif filename.endswith('.pdf'):
                    pdf_bytes = arquivo.file.read()
                    arquivo.file.seek(0)
                    dados = processar_pdf_com_gemini(pdf_bytes, arquivo.filename)
                    dados_consolidados.extend(dados)
                    
                elif filename.endswith(('.jpg', '.jpeg', '.png')):
                    img_bytes = arquivo.file.read()
                    arquivo.file.seek(0)
                    dados = processar_imagem_com_gemini(img_bytes, arquivo.filename)
                    dados_consolidados.extend(dados)
                    
            except Exception as e:
                print(f"[AVISO] Erro ao processar {arquivo.filename}: {e}")
                continue
        
        if not dados_consolidados:
            raise ValueError("Nenhum dado válido foi encontrado nos arquivos enviados.")
        
        # Calcula com a função isolada (sem overrides no primeiro processamento)
        relatorio, preview, totais_semanais = calcular_relatorio(dados_consolidados, settings_dict, status_overrides=None)
        
        if not relatorio:
            raise ValueError("Não foi possível calcular o relatório.")
        
        # Gera Excel PROFISSIONAL - NOVO v6.1: passa totais semanais para códigos 150/200
        arquivo_excel = gerar_excel(relatorio, settings_dict, totais_semanais)
        encoded_file = base64.b64encode(arquivo_excel.getvalue()).decode('utf-8')
        filename = f"Espelho_Ponto_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
        
        return JSONResponse({
            "preview": preview,
            "file": encoded_file,
            "filename": filename
        })
    
    except ValueError as e:
        print(f"❌ ERRO CRÍTICO (ValueError): {e}")
        return JSONResponse({"erro": str(e)}, status_code=400)
    except Exception as e:
        print(f"[ERRO] CRÍTICO: {e}")
        return JSONResponse(
            {"erro": "Ocorreu um erro interno ao processar os cálculos. Por favor, tente novamente mais tarde."}, 
            status_code=500
        )


@app.post("/recalcular")
async def recalcular_com_edicoes(payload: dict):
    """
    Rota: recebe dados editados do frontend e recalcula o Excel.
    
    Alterado para aceitar dict em vez de Pydantic model para robustez máxima.
    """
    warnings: List[str] = []  # Rastreia erros/alertas durante processamento
    
    try:
        print(f"\n{'='*70}")
        print(f"[INFO] Recalculando com dados editados pelo usuário")
        print(f"{'='*70}\n")
        
        # Validação básica do payload
        if not isinstance(payload, dict):
            raise ValueError("Payload deve ser um objeto JSON válido")
        
        dados_corrigidos = payload.get('dados_corrigidos')
        settings = payload.get('configuracoes')
        
        if not dados_corrigidos or not settings:
            raise ValueError("Payload incompleto: faltam 'dados_corrigidos' ou 'configuracoes'")
        
        # Reconstrói a lista de batidas a partir do JSON editado
        dados_reconstruidos = []
        status_overrides = {}
        
        ano_base_detectado = None  # Detectar ano dos dados
        
        # PRIORIDADE 1: Verificar se o campo 'ano' foi enviado explicitamente nas configurações
        if 'ano' in settings and settings['ano']:
            try:
                ano_base_detectado = int(settings['ano'])
                print(f"[OK] Ano base detectado nas configurações: {ano_base_detectado}")
            except (ValueError, TypeError):
                print(f"[AVISO] Campo 'ano' nas configurações é inválido: {settings['ano']}")
                ano_base_detectado = None
        
        # PRIORIDADE 2: Extrair o ano da primeira data válida encontrada nos dados
        if ano_base_detectado is None:
            for func_data in dados_corrigidos.get('preview', []):
                for dia_info in func_data.get('dias', []):
                    dia_mes = dia_info.get('data', '')
                    # Tenta fazer parse direto se há "DD/MM/YYYY"
                    if '/' in dia_mes:
                        partes = dia_mes.split('/')
                        if len(partes) == 3:
                            try:
                                ano_base_detectado = int(partes[2])
                                print(f"[OK] Ano base detectado na primeira data: {ano_base_detectado}")
                                break
                            except ValueError:
                                pass
                    # Busca no histórico de batidas
                    batidas_str = dia_info.get('batidas', '')
                    if '/' in batidas_str:
                        try:
                            partes = batidas_str.split()[0].split('/')
                            if len(partes) == 3:
                                ano_base_detectado = int(partes[2])
                                print(f"[OK] Ano base detectado nas batidas: {ano_base_detectado}")
                                break
                        except (ValueError, IndexError):
                            pass
                if ano_base_detectado:
                    break
        
        # PRIORIDADE 3: Último recurso - usar ano corrente com AVISO no console
        if ano_base_detectado is None:
            ano_base_detectado = datetime.now().year
            print(f"[AVISO] CRÍTICO: Sistema não conseguiu inferir ano base dos dados.")
            print(f"[AVISO] USANDO ANO CORRENTE COMO FALLBACK: {ano_base_detectado}")
            print(f"[AVISO] Pode haver imprecisão em cálculos de DSR e Sábados!")
        
        # Agora processa cada funcionário com o ano_base_detectado
        for func_data in dados_corrigidos.get('preview', []):
            funcionario = func_data['funcionario']
            
            for dia_info in func_data['dias']:
                try:
                    dia_mes = dia_info['data']
                    data_completa = datetime.strptime(f"{dia_mes}/{ano_base_detectado}", "%d/%m/%Y").date()
                    
                    # Captura o status vindo do frontend
                    status_front = dia_info.get('status')
                    if status_front:
                        override_key = f"{funcionario}|{data_completa.isoformat()}"
                        status_overrides[override_key] = status_front
                    
                    # Parse das batidas com NORMALIZAÇÃO
                    # Novo formato pode vir como objeto com 4 campos (Entrada1, Saída1, Entrada2, Saída2)
                    # ou como string compatível com o anterior (ex: "07:30 → 12:00 → 13:00 → 17:30")
                    
                    batidas_str = dia_info.get('batidas', '')
                    batidas_4cols = dia_info.get('batidas_4cols', {})
                    
                    horarios = []
                    
                    # PRIORIDADE 1: Usar objeto de 4 colunas se disponível
                    if batidas_4cols and isinstance(batidas_4cols, dict):
                        for chave in ['entrada_1', 'saida_1', 'entrada_2', 'saida_2']:
                            valor = batidas_4cols.get(chave)
                            hora_norm = normalizar_horario(valor)
                            if hora_norm:
                                horarios.append(datetime.combine(data_completa, hora_norm))
                    
                    # PRIORIDADE 2: Parse da string se batidas_4cols vazio
                    elif batidas_str and batidas_str not in ["Falta", "Sem registro", "Atestado", "Folga", "Feriado"]:
                        horarios_raw = batidas_str.split(' -> ')
                        # Fallback se não encontrar a seta ASCII
                        if len(horarios_raw) == 1 and ' → ' in batidas_str:
                             horarios_raw = batidas_str.split(' → ')
                        
                        for horario_str in horarios_raw:
                            try:
                                horario_str = horario_str.strip()
                                hora_norm = normalizar_horario(horario_str)
                                if hora_norm:
                                    horarios.append(datetime.combine(data_completa, hora_norm))
                            except ValueError as ve:
                                # ROBUSTEZ: Log detalhado de erro de formatação
                                warning_msg = f"[AVISO] Horário inválido para {funcionario} em {data_completa.strftime('%d/%m/%Y')}: '{horario_str}' (erro: {str(ve)})"
                                print(warning_msg)
                                warnings.append(warning_msg)
                                continue
                    
                    # Se conseguiu processar horários, adiciona à lista
                    if horarios:
                        for hor in horarios:
                            dados_reconstruidos.append({
                                "nome": funcionario,
                                "data": data_completa,
                                "hora": hor.time()
                            })
                
                except ValueError as ve:
                    # ROBUSTEZ: Log detalhado de erro ao parsear data
                    warning_msg = f"[AVISO] Data inválida para {funcionario}: '{dia_mes}' (erro: {str(ve)})"
                    print(warning_msg)
                    warnings.append(warning_msg)
                    continue
                except Exception as e:
                    # ROBUSTEZ: Erro genérico com contexto completo
                    warning_msg = f"[AVISO] Erro ao processar dia {dia_info.get('data')} para {funcionario}: {str(e)}"
                    print(warning_msg)
                    warnings.append(warning_msg)
                    continue
        
        print(f"[DATA] Ano base detectado para recálculo: {ano_base_detectado}")
        if warnings:
            print(f"[AVISO] Total de warnings durante processamento: {len(warnings)}")
        
        # Recalcula PASSANDO OS OVERRIDES
        relatorio, preview, totais_semanais = calcular_relatorio(dados_reconstruidos, settings, status_overrides=status_overrides)
        
        if relatorio is None:
            raise ValueError("Não foi possível recalcular.")
        
        # Gera novo Excel PROFISSIONAL - NOVO v6.1: passa totais semanais para códigos 150/200
        arquivo_excel = gerar_excel(relatorio, settings, totais_semanais)
        encoded_file = base64.b64encode(arquivo_excel.getvalue()).decode('utf-8')
        filename = f"Espelho_Recalculado_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
        
        print(f"[OK] Recálculo concluído: {filename}\n")
        
        # ROBUSTEZ: Inclui warnings na resposta JSON para o frontend
        response_data = {
            "preview": preview,
            "file": encoded_file,
            "filename": filename
        }
        
        if warnings:
            response_data["warnings"] = warnings
        
        return JSONResponse(response_data)
    
    except ValueError as e:
        print(f"[ERRO] ERRO CRÍTICO (ValueError): {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            {
                "erro": str(e),
                "warnings": warnings
            }, 
            status_code=400
        )
    except Exception as e:
        print(f"[ERRO] ERRO CRÍTICO: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(
            {
                "erro": "Ocorreu um erro interno ao processar os cálculos. Por favor, tente novamente mais tarde.",
                "detalhes": str(e),
                "warnings": warnings
            }, 
            status_code=500
        )
